import functions
from openpyxl import load_workbook, Workbook
import os
import warnings


# Ignore the "UserWarning: Data Validation extension is not supported and will be removed" on run as we are writing to a different file anyways
warnings.simplefilter(action='ignore', category=UserWarning)    


""" ------------ Read Data From Input Sheet ------------ """
workbook = load_workbook(filename="courseinput.xlsx")
sheets = workbook.sheetnames
datasheet = functions.chooseSheet(workbook, sheets, 1)   # Read from data (second) sheet
courses = functions.readCourses(datasheet)
courses = functions.filterCourselist(courses)

# for course_choices in courses:
#     for Course in course_choices:
#         print(Course.name)
#         print(Course.times)
#         for Time in Course.times:
#             print(Time.day, Time.start, Time.end)
#         print()
#     print()

#print(courses)
schedules = functions.treeCombinations(courses)
#print(schedules)
for schedule in schedules:
    for Course in schedule:
        print(Course.name)
        for Time in Course.times:
            print(Time.day, Time.start, Time.end)
        print()
    print("------------")

""" ---------------------------------------------------- """


""" -------- Make New Sheet and Draw Schedules! -------- """
workbook = Workbook()
workingsheet = workbook.active

functions.drawOutline(workingsheet, len(schedules))
functions.fillSchedules(workingsheet, schedules)
""" ---------------------------------------------------- """


""" ------------ Save and Open Spreadsheet! ------------ """
workbook.save(filename="timetable.xlsx")
os.system("start EXCEL.EXE timetable.xlsx")
os.system("start EXCEL.EXE courseinput.xlsx")
""" ---------------------------------------------------- """
