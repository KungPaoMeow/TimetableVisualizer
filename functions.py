import math
import openpyxl
from typing import List
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Color

TIMES = ["8:30 AM", "9:30 AM", "10:30 AM", "11:30 AM", "12:30 PM", "1:30 PM", "2:30 PM", "3:30 PM", "4:30 PM", "5:30 PM", "6:30 PM", "7:30 PM", "8:30 PM", "9:30 PM"]

COLORS = [PatternFill(start_color='ac92eb', end_color='ac92eb', fill_type='solid'), PatternFill(start_color='4fc1e8', end_color='4fc1e8', fill_type='solid'), 
    PatternFill(start_color='a0d568', end_color='a0d568', fill_type='solid'), PatternFill(start_color='ffce54', end_color='ffce54', fill_type='solid'), PatternFill(start_color='ed5564', end_color='ed5564', fill_type='solid')]

class Course:
   def __init__(self, name):
      self.name = name;
      self.times = []   # Holds Time objects


class Time:
    def __init__(self, day, start, end):
        self.day = day
        self.start = start
        self.end = end


def chooseSheet(workbook, sheets, index) -> openpyxl.worksheet.worksheet.Worksheet:
    return workbook[sheets[index]]


def readCourses(datasheet) -> List[List[Course]]:
    MAX_NUM_COURSES = 5
    TOP_ROW = 3
    MAX_COURSE_OFFERINGS = 2
    courses = []

    for c in range(MAX_NUM_COURSES):
        courses.append([])

    row = TOP_ROW

    for i in range(MAX_COURSE_OFFERINGS):      
        coursecol = "B"
        daycol = "C"
        startcol = "D"
        stopcol = "E"

        for c in range(MAX_NUM_COURSES):  # Columns or courses
            course_name = datasheet[coursecol+str(TOP_ROW)].value
            if (course_name != None):
                courses[c].append(Course(course_name))

            for t in range(5):  # Inside Rows or times
                day = datasheet[daycol+str(row+t)].value
                if (day != None):
                    courses[c][i].times.append(Time(day, datasheet[startcol+str(row+t)].value, datasheet[stopcol+str(row+t)].value))    # Times get read as HH:MM:SS

            # Each col + 5
            coursecol = chr(ord(coursecol)+5)  # Turns into ASCII value to increment and change back to letter
            daycol = chr(ord(daycol)+5)
            startcol = chr(ord(startcol)+5)
            stopcol = chr(ord(stopcol)+5)

        row += 6

    return courses


def filterCourselist(course_list):
    """ Removes courses that have an empty times list, and groups of courses with no courses """
    for group in course_list:
        if len(group) == 0:
            course_list.remove(group)
        else:
            for Course in group:
                if len(Course.times) == 0:
                    group.remove(Course)
    return course_list
    

def treeCombinations(group_list):
    """ Returns all possible combinations from a tree diagram """
    result = []
    result_length = 1
    freq = []   # amount of times each element in group will appear in final list

    for i in range(len(group_list)): 
        result_length *= len(group_list[i]) 

    for i in range(len(group_list)):    # Need result_length for this
        freq.append(result_length/len(group_list[i]))

    for i in range(result_length):
        result.append([])


    cycle_mult = 1
    for i in range(len(group_list)):
        if (i != 0):    # Do nothing on first run of outer loop
            cycle_mult *= len(group_list[i-1]) 
        for j in range(result_length):     
            index = math.floor((j*cycle_mult)/freq[i])
            result[j].append(group_list[i][math.floor(index%len(group_list[i]))])
        #print(cycle_mult)

    #print(freq)
    return result


def drawOutline(workingsheet, num_outlines):
    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    THIN = Side(border_style="thin", color="000000")
    DOUBLE = Side(border_style="double", color="000000")
    CENTER_ALIGNED_TEXT = Alignment(horizontal="center", vertical="center")

    for num in range(num_outlines):
        row = 3 +30*num
        for i in range(14):     # Time column
            top_cell_index = "B"+str(row+i*2)       # Use this to change merged cell value and font settings
            bot_cell_index = "B"+str(row+1+i*2)
            workingsheet.merge_cells(top_cell_index+":"+bot_cell_index)
            workingsheet[top_cell_index] = TIMES[i]
            workingsheet[top_cell_index].number_format = "hh:mm:ss"
            #print(toTwentyfourHour(workingsheet[top_cell_index].value))
            workingsheet[top_cell_index].alignment = CENTER_ALIGNED_TEXT
            workingsheet[top_cell_index].border = Border(top=DOUBLE, left=DOUBLE, right=THIN)
            workingsheet[bot_cell_index].border = Border(left=DOUBLE, right=THIN, bottom=DOUBLE)
        
        for i in range(28):     # Add borders to right side
            cell_index = "G"+str(row+i)
            if (row+i)%30 == 0:
                workingsheet[cell_index].border = Border(right=DOUBLE, bottom=DOUBLE)
            else:
                workingsheet[cell_index].border = Border(right=DOUBLE)

        row = 2 +30*num
        col = "C"
        for i in range(5):      # Days row
            cell_index = col+str(row)
            workingsheet[cell_index] = DAYS[i]
            workingsheet[cell_index].alignment = CENTER_ALIGNED_TEXT
            workingsheet[cell_index].border = Border(top=DOUBLE, left=DOUBLE, right=DOUBLE, bottom=THIN)
            workingsheet.column_dimensions[col].width = 15
            col = chr(ord(col)+1)

        row = 30 +30*num
        col = "C"
        for i in range(4):      # Add borders to bottom
            cell_index = col+str(row)
            workingsheet[cell_index].border = Border(bottom=DOUBLE)
            col = chr(ord(col)+1)


def toTwentyfourHour(twelve_hour):
    if (twelve_hour[-2:] == "PM" and twelve_hour[:2] == str(12)):
        return twelve_hour[:5] + ":00"
    elif (twelve_hour[-2:] == "PM" and twelve_hour[:2] != str(12)):
        if len(twelve_hour) < 8:
            return str(int(twelve_hour[:1])+12) + twelve_hour[1:4] + ":00"
        else:
            return str(int(twelve_hour[:2])+12) + twelve_hour[2:5] + ":00"
    elif len(twelve_hour) < 8:      # AMs
        return "0" + twelve_hour[:-3] + ":00"
    else:       # AMs
        return twelve_hour[:-3] + ":00"
         

def fillSchedules(workingsheet, schedules):
    #lightPurple = PatternFill(start_color='D8BFD8', end_color='D8BFD8', fill_type='solid')
    out_row = 2
    in_row = 4
    for schedule in schedules:
        color = -1
        for Course in schedule:
            color += 1
            for Time in Course.times:
                # Find the day
                col = "C"
                for c in range(5):
                    if (Time.day == workingsheet[col + str(out_row)].value):
                        # Find start and end time
                            for t in range(len(TIMES)):
                                #print("`"+toTwentyfourHour(t)+"` `"+str(Time.start))
                                if (toTwentyfourHour(TIMES[t])[:-3] == str(Time.start)[:-3]):       # ex 10:30 start
                                    adjust = 0
                                    if (str(Time.end)[-5:-3] == "00"):      # # adjustment for xx:00 end
                                        adjust = 1

                                    for x in range((int(str(Time.end)[:-6]) - int(str(Time.start)[:-6])) *2 - adjust):   # Fill in x squares from start to stop time
                                        if (workingsheet[col + str(in_row+t*2 +x)].value == None):
                                            workingsheet[col + str(in_row+t*2 +x)].value = " "+Course.name
                                        else:
                                            workingsheet[col + str(in_row+t*2 +x)].value += " "+Course.name
                                        workingsheet[col + str(in_row+t*2 +x)].fill = COLORS[color]

                                elif (toTwentyfourHour(TIMES[t])[:-6] == str(Time.start)[:-6]):     # ex 10:00 start         make it work
                                    adjust = 0
                                    if (str(Time.end)[-5:-3] == "30"):      # adjustment for xx:30 end
                                        adjust = 1

                                    for x in range((int(str(Time.end)[:-6]) - int(str(Time.start)[:-6])) *2 + adjust):   # Fill in x squares from start to stop time
                                        if (workingsheet[col + str(in_row+t*2 +x -1)].value == None):
                                            workingsheet[col + str(in_row+t*2 +x -1)].value = " "+Course.name
                                        else:
                                            workingsheet[col + str(in_row+t*2 +x -1)].value += " "+Course.name
                                        workingsheet[col + str(in_row+t*2 +x -1)].fill = COLORS[color]

                    col = chr(ord(col)+1)
        out_row += 30
        in_row += 30


# print(int("10")-int("09"))
# print(toTwentyfourHour("8:30 AM")[-5:-3])
# print(float(("1,570.00").replace(",", "")))